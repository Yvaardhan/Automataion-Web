import streamlit as st
import re
import os
import csv
import time
import glob
import pandas as pd
from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.edge.options import Options
from selenium.webdriver.support import expected_conditions as EC
import tempfile
import subprocess
import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Constants
edge_options = Options()
edge_options.add_argument("--headless")
edge_options.add_argument("--disable-gpu")
edge_options.add_argument("--no-sandbox")
edge_options.add_argument("--disable-dev-shm-usage")

driver_path = r"C:\Users\yash.v1\Documents\Web Scapping Project\edgedriver_win64\msedgedriver.exe"
download_dir = r"C:\Users\yash.v1\Documents\Web Scapping Project"
website = "http://107.108.175.239:8000/DashBoard/dataPage"

# Dictionary to store Model_name (Model ID key from system) and its corresponding display name
model_dict = {
    # T09 Models - 2023 Series
    "23_OSCP_GMT9_T09": "G97NC_T09",
    "23_OSCS_GMT9_T09": "G95SC/SD_T09",
    "23_OSCS_SMT9_T09": "S90PC_T09",
    "23_PTML_SMT7_T09": "M70C_T09",
    "23_PTML_SMT8_T09": "M80C_T09",
    "23_NKL_SMT5_T09": "M50C_T09",
    
    # T10 Models - 2023 Series
    "23_OSCP_GM9_T10": "G97NC_T10",
    "23_OSCS_GM9_T10": "G95SC/SD_T10",
    "23_OSCS_SM9_T10": "S90PC_T10",
    "23_PTML_SM7_T10": "M70C_T10",
    "23_PTML_SM8_T10": "M80C_T10",
    "23_NKL_SM5_T10": "M50C_T10",
    
    # T09 Models - 2025 Series
    "25_PTM_MSC": "LSM7F_T09",
    "25_PTM_SMT8": "M80F_T09",
    "25_RSL_SMT7": "M70F_T09",
    "25_RSM_SMT9": "M90SF_T09",
    "25_RSP_SM9": "M90SF_P_T09",
    "25_RSSF_SMT5": "M50F_T09",
    
    # T10 Models - 2025 Series
    "25_PTM_MSC_T10": "LSM7F_T10",
    "25_PTM_SM8_T10": "M80F_T10",
    "25_RSL_SM7_T10": "M70F_T10",
    "25_RSM_SM9_T10": "M90SF_T10",
    "25_RSP_SM9_T10": "M90SF_P_T10",
    "25_RSSF_SM5_T10": "M50F_T10",
    
    # T09 Models - 2024 Series
    "24_NKL_SM5_T09": "M50D_T09",
    "24_NKL_SMT7_T09": "M70D_T09,M70DO_T09,M1ED_T09,M1EDO_T09",
    "24_NKL_SMT7_AT_T09": "M70D_AT_T09",
    "24_PTML_GMT7_T09": "G70D_T09",
    "24_PTML_GMT7_AT_T09": "G70D_AT_T09",
    "24_PTML_GMT8_T09": "G85SD_T09",
    "24_PTML_GMT8_AT_T09": "G85SD_AT_T09",
    "24_PTML_SMT8_T09": "M80D_T09",
    "24_PTML_SMT8_AT_T09": "M80D_AT_T09",
    
    # T10 Models - 2024 Series
    "24_NKL_SM5_T10": "M50D_T10",
    "24_NKL_SM7_T10": "M70D_T10,M70DO_T10,M1ED_T10,M1EDO_T10",
    "24_NKL_SM7_T10_AT": "M70D_AT_T10",
    "24_PTML_GM7_T10": "G70D_T10",
    "24_PTML_GM7_T10_AT": "G70D_AT_T10",
    "24_PTML_GM8_T10": "G85SD_T10",
    "24_PTML_GM8_T10_AT": "G85SD_AT_T10",
    "24_PTML_SM8_T10": "M80D_T10",
    "24_PTML_SM8_T10_AT": "M80D_AT_T10",
}

exception_app_names = {
    "User Guide",
    "E_manual",
}

def extract_name(full_name):
    pattern = r'\d+(?:[A-Za-z0-9_ ]+)\s*\[PRD\]|\d+(?:[A-Za-z0-9_ ]+)\s*\[STG\]'
    match = re.search(pattern, full_name)
    return match.group(0) if match else "No match found."

def compute_state_value(master_df):
    """
    Compute the State Value column for the Master Excel DataFrame.
    
    Args:
        master_df (pd.DataFrame): DataFrame containing the Master Excel data
    
    Returns:
        pd.DataFrame: Updated DataFrame with State Value column
    """
    # Initialize State Value column
    master_df['State Value'] = None
    
    # Get version columns (all columns except "App Name")
    version_columns = [col for col in master_df.columns if col != 'App Name' and col != 'State Value']
    
    print(f"Found {len(version_columns)} version columns for State Value computation")
    
    # Process each row
    for index, row in master_df.iterrows():
        app_name = str(row.get('App Name', '')).strip()
        
        # Get all version values for this row
        version_values = []
        has_none = False
        
        for col in version_columns:
            value = row.get(col)
            # Check for NONE, Not Found, or empty values
            if pd.isna(value) or str(value).strip().upper() == 'NONE' or str(value).strip() == '' or str(value).strip() == 'Not Found':
                has_none = True
            else:
                version_values.append(str(value).strip())
        
        # Condition 2: If the App Name is Started with "unified-drm" then State Value = 1
        # This takes precedence over all other conditions
        if app_name.startswith('unified-drm'):
            master_df.at[index, 'State Value'] = 1
            print(f"Row {index}: App '{app_name}' starts with 'unified-drm' -> State Value = 1")
            continue
        
        # Condition 1: If all the App Versions are same in the same row then State Value = 0
        if version_values and len(version_values) > 0:
            unique_versions = list(set(version_values))
            if len(unique_versions) == 1 and not has_none:
                master_df.at[index, 'State Value'] = 0
                print(f"Row {index}: All versions are the same ('{unique_versions[0]}') -> State Value = 0")
                continue
        
        # Condition 3: If there is any NONE value in a particular row apart from above conditions then State Value = 2.1
        if has_none:
            master_df.at[index, 'State Value'] = 2.1
            print(f"Row {index}: Found NONE value in row -> State Value = 2.1")
            continue
        
        # Condition 4: If in a particular row any App version is different then State Value = 2.2
        if version_values and len(version_values) > 0:
            unique_versions = list(set(version_values))
            if len(unique_versions) > 1:
                master_df.at[index, 'State Value'] = 2.2
                print(f"Row {index}: Different versions found: {unique_versions} -> State Value = 2.2")
            else:
                # All versions are the same but with NONE values
                master_df.at[index, 'State Value'] = 0
                print(f"Row {index}: All versions are the same (with NONE) -> State Value = 0")
        else:
            # Default case: no valid version data
            master_df.at[index, 'State Value'] = 3
            print(f"Row {index}: No valid version data -> State Value = 0")
    
    return master_df


def apply_row_colors(excel_file_path):
    """
    Apply row colors based on State Value.
    
    Colors:
    - 0: Medium Green
    - 1: Gray
    - 2.1: Yellow
    - 2.2: Orange
    - 3: Medium Red
    
    Args:
        excel_file_path (str): Path to the Excel file
    """
    # Define color fills
    colors = {
        0.0: PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),      # Medium Green
        1.0: PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"),      # Gray
        2.1: PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),      # Yellow
        2.2: PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),      # Orange
        3.0: PatternFill(start_color="CD5C5C", end_color="CD5C5C", fill_type="solid")       # Medium Red
    }
    
    # Load the workbook
    wb = load_workbook(excel_file_path)
    ws = wb.active
    
    # Find the State Value column index
    state_value_col = None
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value == "State Value":
            state_value_col = col_idx
            break
    
    if state_value_col is None:
        print("Warning: State Value column not found")
        return
    
    # Apply colors to each row based on State Value
    for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
        state_value = ws.cell(row=row_idx, column=state_value_col).value
        
        if state_value in colors:
            fill = colors[state_value]
            # Apply fill to all cells in the row
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill
    
    # Save the workbook
    wb.save(excel_file_path)
    print(f"✅ Row colors applied based on State Value!")


def get_row_color(state_value):
    """
    Get background color based on State Value.
    
    Colors:
    - 0: Medium Green
    - 1: Gray
    - 2.1: Yellow
    - 2.2: Orange
    - 3: Medium Red
    """
    colors = {
        0.0: "#90EE90",  # Medium Green
        1.0: "#D3D3D3",  # Gray
        2.1: "#FFFF00",  # Yellow
        2.2: "#FFA500",  # Orange
        3.0: "#CD5C5C"   # Medium Red
    }
    return colors.get(state_value, "#FFFFFF")


def style_dataframe(df):
    """
    Apply color styling to the dataframe based on State Value.
    """
    # Create a copy to avoid modifying original
    df_display = df.copy()
    
    # Format State Value column to show 2.1 instead of 2.1000
    if 'State Value' in df_display.columns:
        df_display['State Value'] = df_display['State Value'].apply(
            lambda x: f"{x:.1f}" if pd.notna(x) else x
        )
    
    def highlight_row(row):
        # Get the original state value for color matching
        state_value = float(row.get('State Value')) if row.get('State Value') else None
        color = get_row_color(state_value)
        return [f'background-color: {color}; color: #2C3E50; font-weight: 400' for _ in row]
    
    return df_display.style.apply(highlight_row, axis=1)


def run_script(tsv_path, username, key):
    with tempfile.TemporaryDirectory() as temp_dir:
        options = webdriver.EdgeOptions()
        prefs = {
            "download.default_directory": temp_dir,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True,
        }
        options.add_experimental_option("prefs", prefs)

        service = Service(executable_path=driver_path)
        driver = webdriver.Edge(service=service, options=options)
        wait = WebDriverWait(driver, 20)

        # Login
        driver.get(website)
        wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="id_username"]'))).send_keys(username)
        wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="id_password"]'))).send_keys(key)
        wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div/div/form/div/button'))).click()
        time.sleep(2)

        Comparison_button_xpath = '/html/body/div[3]/ul/li[4]/label/span'
        wait.until(EC.element_to_be_clickable((By.XPATH, Comparison_button_xpath))).click()
        time.sleep(3)

        output_data = [] 
        unique_app_names = set()

        with open(tsv_path, newline='', encoding='utf-8') as tsvfile:
            reader = csv.DictReader(tsvfile, delimiter='\t')
            for row in reader:
                Reference_Server = row['Reference_Server']
                Current_Server = row['Current_Server']
                Model_id_reference = row['Model_id_reference']
                Model_id_current = row['Model_id_current']
                Country_Reference = row['Country_reference']
                Country_Current = row['Country_current']
                Infolink_version_reference = row['Infolink_version_reference']
                Infolink_version_current = row['Infolink_version_current']

                # Fill form
                ref_server_xpath = '//*[@id="id_Server_Type_Reference"]'
                driver.execute_script("arguments[0].click();", wait.until(EC.element_to_be_clickable((By.XPATH, ref_server_xpath))))
                driver.find_element(By.XPATH, f"{ref_server_xpath}/option[1]" if Reference_Server.lower() == "staging" else f"{ref_server_xpath}/option[2]").click()

                cur_server_xpath = '//*[@id="id_Server_Type_Current"]'
                driver.execute_script("arguments[0].click();", wait.until(EC.element_to_be_clickable((By.XPATH, cur_server_xpath))))
                driver.find_element(By.XPATH, f"{cur_server_xpath}/option[1]" if Current_Server.lower() == "staging" else f"{cur_server_xpath}/option[2]").click()

                driver.find_element(By.XPATH, '//*[@id="id_Model_ID_Reference"]').clear()
                driver.find_element(By.XPATH, '//*[@id="id_Model_ID_Reference"]').send_keys(Model_id_reference)

                driver.find_element(By.XPATH, '//*[@id="id_Model_ID_Current"]').clear()
                driver.find_element(By.XPATH, '//*[@id="id_Model_ID_Current"]').send_keys(Model_id_current)

                driver.find_element(By.XPATH, '//*[@id="id_Country_Reference"]').clear()
                driver.find_element(By.XPATH, '//*[@id="id_Country_Reference"]').send_keys(Country_Reference)

                driver.find_element(By.XPATH, '//*[@id="id_Country_Current"]').clear()
                driver.find_element(By.XPATH, '//*[@id="id_Country_Current"]').send_keys(Country_Current)

                driver.find_element(By.XPATH, '//*[@id="id_InfoLink_Version_Reference"]').clear()
                driver.find_element(By.XPATH, '//*[@id="id_InfoLink_Version_Reference"]').send_keys(Infolink_version_reference)

                driver.find_element(By.XPATH, '//*[@id="id_InfoLink_Version_Current"]').clear()
                driver.find_element(By.XPATH, '//*[@id="id_InfoLink_Version_Current"]').send_keys(Infolink_version_current)

                driver.find_element(By.XPATH, '//*[@id="ComparissionForm"]/div[5]/center').click()
                time.sleep(5)

                # Download CSV
                try:
                    download_button = wait.until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="ComparissionTable_wrapper"]/div[1]/button[3]'))
                    )
                    driver.execute_script("arguments[0].click();", download_button)
                except Exception as e:
                    print(f"Download button error: {e}")
                    driver.quit()
                    continue 

                time.sleep(10)
                list_of_files = glob.glob(os.path.join(temp_dir, '*.csv'))
                if not list_of_files:
                    print("Error: No CSV file was downloaded.")
                    driver.quit()
                    return

                latest_file = max(list_of_files, key=os.path.getctime)
                new_filename = f"{Model_id_reference}_vs_{Model_id_current}.csv"
                new_filepath_in_temp = os.path.join(temp_dir, new_filename) 
                if os.path.basename(latest_file) != new_filename:
                    os.rename(latest_file, new_filepath_in_temp)

                output_data.append({
                    "Reference_Model": Model_id_reference,
                    "Current_Model": Model_id_current,
                    "Downloaded_File_Path": new_filepath_in_temp,
                })
                print(f"Downloaded and Saved (Temp): {new_filepath_in_temp}")
                time.sleep(10)
                # Collect app names
                with open(new_filepath_in_temp, mode='r', newline='', encoding='utf-8') as file:
                    reader = csv.DictReader(file)
                    for r in reader:
                        app_owner_value = r.get("App Owner", "").strip()
                        app_name = r.get("App Name")
                        if app_owner_value != "None" and app_owner_value or any(app_name.startswith(exception) for exception in exception_app_names):
                            unique_app_names.add(app_name)

        # Create Master Excel
        app_names_df = pd.DataFrame(unique_app_names, columns=["App Name"])
        master_excel_file = os.path.join(download_dir, "master_Excel.xlsx")
        app_names_df.to_excel(master_excel_file, index=False)
        print(f"Master Excel file created at: {master_excel_file}")

        master_df = pd.read_excel(master_excel_file)

        # ✅ Process CSVs while still in temp_dir
        for data_entry in output_data:
            csv_filepath = data_entry["Downloaded_File_Path"]
            csv_df = pd.read_csv(csv_filepath)

            column1 = extract_name(csv_df.columns[1])
            column2 = extract_name(csv_df.columns[2])
            model_id1 = csv_df.columns[1].split()[5]
            model_id2 = csv_df.columns[2].split()[5]
            model_name1 = model_dict[model_id1]
            model_name2 = model_dict[model_id2]
            new_column1 = f"{column1}_Reference_Model \n {model_name1}"
            new_column2 = f"{column2}_Current_Model \n {model_name2}"

            for app_name in master_df["App Name"]:
                matching_rows = csv_df[csv_df["App Name"] == app_name]
                if not matching_rows.empty:
                    extracted_data = matching_rows.iloc[:, 1:3]
                    master_df.loc[master_df["App Name"] == app_name, [new_column1,new_column2]] = extracted_data.iloc[:, 0:2].values
                else:
                    master_df.loc[master_df["App Name"] == app_name, [new_column1,new_column2]] = ["Not Found", "Not Found"]

        reference_columns = [col for col in master_df.columns if "_Reference_Model" in col]
        current_columns = [col for col in master_df.columns if "_Current_Model" in col]
        other_columns = [col for col in master_df.columns if col not in reference_columns + current_columns]
        new_order = other_columns + reference_columns + current_columns
        master_df = master_df[new_order]

        # Compute State Value column
        print("\nComputing State Value column...")
        master_df = compute_state_value(master_df)
        
        # Display summary statistics
        state_counts = master_df['State Value'].value_counts().sort_index()
        print("\nState Value Summary:")
        for state, count in state_counts.items():
            print(f"  State Value {state}: {count} rows")

        master_df.to_excel(master_excel_file, index=False)
        print(f"Master Excel file updated with State Value: {master_excel_file}")

        # Apply row colors based on State Value
        print("\nApplying row colors...")
        apply_row_colors(master_excel_file)

        subprocess.Popen([master_excel_file], shell=True)

        driver.quit()

    return True, master_df, master_excel_file

# --------- Streamlit UI ---------
# Hardcoded credentials for website login
USERNAME = "Yash"
PASSWORD = "Yash@2003"

# Configure page
st.set_page_config(page_title="DNA Automation", page_icon="🧬", layout="wide")

# Custom CSS for simple, clean, light theme
st.markdown("""
    <style>
    .main {
        background-color: #FAFBFC;
    }
    .stApp {
        background-color: #FFFFFF;
    }
    h1, h2, h3, p, span, div, label {
        color: #2C3E50 !important;
    }
    .stButton>button {
        background-color: #4A90E2;
        color: white;
        font-weight: 500;
        padding: 0.6rem 1.5rem;
        border-radius: 6px;
        border: none;
    }
    .stButton>button:hover {
        background-color: #357ABD;
    }
    [data-testid="stMetricValue"] {
        color: #2C3E50 !important;
    }
    </style>
    """, unsafe_allow_html=True)

# Main UI
st.title("🧬 DNA Automation")
st.info(f"📁 Logged in as: **{USERNAME}**")
st.markdown("---")

# File upload
st.subheader("Upload TSV File")
uploaded_file = st.file_uploader("Choose a TSV file", type=["tsv"])

if uploaded_file is not None:
    # Save uploaded file temporarily
    temp_tsv_path = os.path.join(tempfile.gettempdir(), uploaded_file.name)
    with open(temp_tsv_path, "wb") as f:
        f.write(uploaded_file.getbuffer())
    
    st.success(f"✅ File uploaded: {uploaded_file.name}")
    
    # Run automation button
    if st.button("Extract DNA DATA", type="primary"):
        with st.spinner("Running automation... This may take several minutes."):
            try:
                # Using hardcoded credentials for website login
                success, master_df, master_excel_file = run_script(temp_tsv_path, USERNAME, PASSWORD)
                if success:
                    st.success(f"✅ All files downloaded and Master Excel created at: {master_excel_file}")
                    st.balloons()
                    
                    # Store in session state
                    st.session_state['master_df'] = master_df
                    st.session_state['master_excel_file'] = master_excel_file
            except Exception as e:
                st.error(f"❌ An error occurred: {str(e)}")
        
        # Clean up temp file
        if os.path.exists(temp_tsv_path):
            os.remove(temp_tsv_path)

# Display dataframe if it exists in session state
if 'master_df' in st.session_state:
    master_df = st.session_state['master_df']
    
    st.markdown("---")
    
    # DNA App Data Result Analysis
    st.subheader("📊 DNA App Data Result")
    
    # Calculate counts for each state
    state_0_count = len(master_df[master_df['State Value'] == 0])
    state_1_count = len(master_df[master_df['State Value'] == 1])
    state_21_count = len(master_df[master_df['State Value'] == 2.1])
    state_22_count = len(master_df[master_df['State Value'] == 2.2])
    
    # Display analysis on left side
    st.markdown(f'<div style="background-color: #90EE90; padding: 10px; margin: 5px 0; border-radius: 4px; color: #2C3E50; border: 1px solid #E0E0E0;"><b>Count of All Apps with same version:</b> {state_0_count}</div>', unsafe_allow_html=True)
    st.markdown(f'<div style="background-color: #D3D3D3; padding: 10px; margin: 5px 0; border-radius: 4px; color: #2C3E50; border: 1px solid #E0E0E0;"><b>Count of All Drm Apps:</b> {state_1_count}</div>', unsafe_allow_html=True)
    st.markdown(f'<div style="background-color: #FFFF00; padding: 10px; margin: 5px 0; border-radius: 4px; color: #2C3E50; border: 1px solid #E0E0E0;"><b>Count of All Apps with No Version:</b> {state_21_count}</div>', unsafe_allow_html=True)
    st.markdown(f'<div style="background-color: #FFA500; padding: 10px; margin: 5px 0; border-radius: 4px; color: #2C3E50; border: 1px solid #E0E0E0;"><b>Count of All Apps with Version Difference:</b> {state_22_count}</div>', unsafe_allow_html=True)
    
    st.markdown("---")
    
    # Display styled dataframe with all rows and columns
    st.subheader("📋 Complete Data Table")
    st.write(f"Showing all {len(master_df)} rows and {len(master_df.columns)} columns")
    styled_df = style_dataframe(master_df)
    
    # Configure pandas display options to show all rows and columns
    pd.set_option('display.max_rows', None)
    pd.set_option('display.max_columns', None)
    
    # Display with scrolling enabled for all data
    st.dataframe(
        styled_df, 
        height=800,
        hide_index=False,
        column_config=None
    )
