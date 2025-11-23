from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import re
import os
import csv
import time
import glob
import pandas as pd
from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.edge.options import Options
from selenium.webdriver.support import expected_conditions as EC
import tempfile
import subprocess
import traceback
import mysql.connector
from mysql.connector import Error
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import io
import json
import platform

app = Flask(__name__)
CORS(app)  # Enable CORS for React frontend

# Store the latest processed data in memory
latest_data = {
    'master_df': None,
    'excel_bytes': None,
    'statistics': None
}

# Constants
driver_path = r"C:\Users\yash.v1\Documents\Web Scapping Project\edgedriver_win64\msedgedriver.exe"
download_dir = r"C:\Users\yash.v1\Documents\Web Scapping Project"
website = "http://107.108.175.239:8000/DashBoard/dataPage"

# Hardcoded credentials
USERNAME = "Yash"
PASSWORD = "Yash@2003"

# Database credentials
DB_HOST = "localhost"
DB_USER = "root"
DB_PASSWORD = "root"
DB_NAME = "dna_automation"

# Dictionary to store Model_name (Model ID key from system) and its corresponding display name
model_dict = {
    # T09 Models - 2023 Series
    "23_OSCP_GMT9_T09": "G97NC_T09",
    "23_OSCS_GMT9_T09": "G95SC/SD_T09",
    "23_OSCS_SMT9_T09": "S90PC_T09",
    "23_PTML_SMT7_T09": "M70C_T09",
    "23_PTML_SMT8_T09": "M80C_T09",
    "23_NKL_SMT5_T09": "M50C_T09",
    
    # T10 Models - 2023 Series
    "23_OSCP_GM9_T10": "G97NC_T10",
    "23_OSCS_GM9_T10": "G95SC/SD_T10",
    "23_OSCS_SM9_T10": "S90PC_T10",
    "23_PTML_SM7_T10": "M70C_T10",
    "23_PTML_SM8_T10": "M80C_T10",
    "23_NKL_SM5_T10": "M50C_T10",
    
    # T09 Models - 2025 Series
    "25_PTM_MSC": "LSM7F_T09",
    "25_PTM_SMT8": "M80F_T09",
    "25_RSL_SMT7": "M70F_T09",
    "25_RSM_SMT9": "M90SF_T09",
    "25_RSP_SM9": "M90SF_P_T09",
    "25_RSSF_SMT5": "M50F_T09",
    
    # T10 Models - 2025 Series
    "25_PTM_MSC_T10": "LSM7F_T10",
    "25_PTM_SM8_T10": "M80F_T10",
    "25_RSL_SM7_T10": "M70F_T10",
    "25_RSM_SM9_T10": "M90SF_T10",
    "25_RSP_SM9_T10": "M90SF_P_T10",
    "25_RSSF_SM5_T10": "M50F_T10",
    
    # T09 Models - 2024 Series
    "24_NKL_SM5_T09": "M50D_T09",
    "24_NKL_SMT7_T09": "M70D_T09,M70DO_T09,M1ED_T09,M1EDO_T09",
    "24_NKL_SMT7_AT_T09": "M70D_AT_T09",
    "24_PTML_GMT7_T09": "G70D_T09",
    "24_PTML_GMT7_AT_T09": "G70D_AT_T09",
    "24_PTML_GMT8_T09": "G85SD_T09",
    "24_PTML_GMT8_AT_T09": "G85SD_AT_T09",
    "24_PTML_SMT8_T09": "M80D_T09",
    "24_PTML_SMT8_AT_T09": "M80D_AT_T09",
    
    # T10 Models - 2024 Series
    "24_NKL_SM5_T10": "M50D_T10",
    "24_NKL_SM7_T10": "M70D_T10,M70DO_T10,M1ED_T10,M1EDO_T10",
    "24_NKL_SM7_T10_AT": "M70D_AT_T10",
    "24_PTML_GM7_T10": "G70D_T10",
    "24_PTML_GM7_T10_AT": "G70D_AT_T10",
    "24_PTML_GM8_T10": "G85SD_T10",
    "24_PTML_GM8_T10_AT": "G85SD_AT_T10",
    "24_PTML_SM8_T10": "M80D_T10",
    "24_PTML_SM8_T10_AT": "M80D_AT_T10",
}

exception_app_names = {
    "User Guide",
    "E_manual",
}

# ========== Database Helper Functions ==========

def get_db_connection():
    """Create and return a database connection."""
    try:
        connection = mysql.connector.connect(
            host=DB_HOST,
            user=DB_USER,
            password=DB_PASSWORD,
            database=DB_NAME
        )
        return connection
    except Error as e:
        print(f"Error connecting to database: {e}")
        return None

def get_all_model_names():
    """Fetch all model names from the database."""
    connection = get_db_connection()
    if not connection:
        return []
    
    try:
        cursor = connection.cursor()
        cursor.execute("SELECT DISTINCT model_name FROM model_names ORDER BY model_name")
        results = cursor.fetchall()
        model_names = [row[0] for row in results]
        return model_names
    except Error as e:
        print(f"Error fetching model names: {e}")
        return []
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()

def get_model_id_from_name(model_name):
    """Get model ID for a given model name."""
    connection = get_db_connection()
    if not connection:
        return None
    
    try:
        cursor = connection.cursor()
        cursor.execute("SELECT model_id FROM model_names WHERE model_name = %s LIMIT 1", (model_name,))
        result = cursor.fetchone()
        return result[0] if result else None
    except Error as e:
        print(f"Error fetching model ID for {model_name}: {e}")
        return None
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()

def get_all_infolink_servers(server_type=None, active_only=True):
    """
    Fetch all InfoLink servers from the database.
    
    Args:
        server_type (str, optional): Filter by server type ('current', 'reference', 'both')
        active_only (bool): If True, only return active servers
    
    Returns:
        list: List of InfoLink server names
    """
    connection = get_db_connection()
    if not connection:
        return []
    
    try:
        cursor = connection.cursor()
        
        # Build query based on filters
        query = "SELECT DISTINCT server_name FROM infolink_servers WHERE 1=1"
        params = []
        
        if active_only:
            query += " AND is_active = %s"
            params.append(True)
        
        if server_type and server_type in ['current', 'reference', 'both']:
            query += " AND (server_type = %s OR server_type = 'both')"
            params.append(server_type)
        
        query += " ORDER BY server_name DESC"
        
        cursor.execute(query, tuple(params))
        results = cursor.fetchall()
        servers = [row[0] for row in results]
        return servers
    except Error as e:
        print(f"Error fetching InfoLink servers: {e}")
        return []
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()

def get_all_countries(country_type=None, active_only=True):
    """
    Fetch all countries from the database.
    
    Args:
        country_type (str, optional): Filter by country type ('current', 'reference', 'both')
        active_only (bool): If True, only return active countries
    
    Returns:
        list: List of country names
    """
    connection = get_db_connection()
    if not connection:
        return []
    
    try:
        cursor = connection.cursor()
        
        # Build query based on filters
        query = "SELECT DISTINCT country_name FROM countries WHERE 1=1"
        params = []
        
        if active_only:
            query += " AND is_active = %s"
            params.append(True)
        
        if country_type and country_type in ['current', 'reference', 'both']:
            query += " AND (country_type = %s OR country_type = 'both')"
            params.append(country_type)
        
        query += " ORDER BY country_name ASC"
        
        cursor.execute(query, tuple(params))
        results = cursor.fetchall()
        countries = [row[0] for row in results]
        return countries
    except Error as e:
        print(f"Error fetching countries: {e}")
        return []
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()

# ========== End Database Functions ==========

def extract_name(full_name):
    """Extract the pattern from full name."""
    pattern = r'\d+(?:[A-Za-z0-9_ ]+)\s*\[PRD\]|\d+(?:[A-Za-z0-9_ ]+)\s*\[STG\]'
    match = re.search(pattern, full_name)
    return match.group(0) if match else "No match found."

def compute_state_value(master_df):
    """
    Compute the State Value column for the Master Excel DataFrame.
    
    Args:
        master_df (pd.DataFrame): DataFrame containing the Master Excel data
    
    Returns:
        pd.DataFrame: Updated DataFrame with State Value column
    """
    # Initialize State Value column
    master_df['State Value'] = None
    
    # Get version columns (all columns except "App Name")
    version_columns = [col for col in master_df.columns if col != 'App Name' and col != 'State Value']
    
    print(f"Found {len(version_columns)} version columns for State Value computation")
    
    # Process each row
    for index, row in master_df.iterrows():
        app_name = str(row.get('App Name', '')).strip()
        
        # Get all version values for this row
        version_values = []
        has_none = False
        
        for col in version_columns:
            value = row.get(col)
            # Check for NONE, Not Found, or empty values
            if pd.isna(value) or str(value).strip().upper() == 'NONE' or str(value).strip() == '' or str(value).strip() == 'Not Found':
                has_none = True
            else:
                version_values.append(str(value).strip())
        
        # Condition 2: If the App Name is Started with "unified-drm" then State Value = 1
        # This takes precedence over all other conditions
        if app_name.startswith('unified-drm'):
            master_df.at[index, 'State Value'] = 1
            print(f"Row {index}: App '{app_name}' starts with 'unified-drm' -> State Value = 1")
            continue
        
        # Condition 1: If all the App Versions are same in the same row then State Value = 0
        if version_values and len(version_values) > 0:
            unique_versions = list(set(version_values))
            if len(unique_versions) == 1 and not has_none:
                master_df.at[index, 'State Value'] = 0
                print(f"Row {index}: All versions are the same ('{unique_versions[0]}') -> State Value = 0")
                continue
        
        # Condition 3: If there is any NONE value in a particular row apart from above conditions then State Value = 2.1
        if has_none:
            master_df.at[index, 'State Value'] = 2.1
            print(f"Row {index}: Found NONE value in row -> State Value = 2.1")
            continue
        
        # Condition 4: If in a particular row any App version is different then State Value = 2.2
        if version_values and len(version_values) > 0:
            unique_versions = list(set(version_values))
            if len(unique_versions) > 1:
                master_df.at[index, 'State Value'] = 2.2
                print(f"Row {index}: Different versions found: {unique_versions} -> State Value = 2.2")
            else:
                # All versions are the same but with NONE values
                master_df.at[index, 'State Value'] = 0
                print(f"Row {index}: All versions are the same (with NONE) -> State Value = 0")
        else:
            # Default case: no valid version data
            master_df.at[index, 'State Value'] = 3
            print(f"Row {index}: No valid version data -> State Value = 0")
    
    return master_df


def get_row_color(state_value):
    """
    Get background color based on State Value.
    
    Colors:
    - 0: Medium Green
    - 1: Gray
    - 2.1: Yellow
    - 2.2: Orange
    - 3: Medium Red
    """
    colors = {
        0.0: "#90EE90",  # Medium Green
        1.0: "#D3D3D3",  # Gray
        2.1: "#FFFF00",  # Yellow
        2.2: "#FFA500",  # Orange
        3.0: "#CD5C5C"   # Medium Red
    }
    return colors.get(state_value, "#FFFFFF")


def apply_row_colors(excel_file_path):
    """
    Apply row colors based on State Value.
    
    Colors:
    - 0: Medium Green
    - 1: Gray
    - 2.1: Yellow
    - 2.2: Orange
    - 3: Medium Red
    
    Args:
        excel_file_path (str): Path to the Excel file
    """
    # Define color fills
    colors = {
        0.0: PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),      # Medium Green
        1.0: PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"),      # Gray
        2.1: PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),      # Yellow
        2.2: PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),      # Orange
        3.0: PatternFill(start_color="CD5C5C", end_color="CD5C5C", fill_type="solid")       # Medium Red
    }
    
    # Load the workbook
    wb = load_workbook(excel_file_path)
    ws = wb.active
    
    # Find the State Value column index
    state_value_col = None
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value == "State Value":
            state_value_col = col_idx
            break
    
    if state_value_col is None:
        print("Warning: State Value column not found")
        return
    
    # Apply colors to each row based on State Value
    for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
        state_value = ws.cell(row=row_idx, column=state_value_col).value
        
        if state_value in colors:
            fill = colors[state_value]
            # Apply fill to all cells in the row
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill
    
    # Save the workbook
    wb.save(excel_file_path)
    print(f"✅ Row colors applied based on State Value!")


def run_automation(rows_data):
    """
    Main automation function that processes the input rows.
    
    Args:
        rows_data: List of dictionaries containing comparison data
        
    Returns:
        Boolean indicating success
    """
    # Verify EdgeDriver exists
    if not os.path.exists(driver_path):
        current_os = platform.system()
        error_msg = f"""❌ EdgeDriver not found at: {driver_path}

Current OS: {current_os}
Expected OS: Windows

⚠️ This application requires Microsoft Edge WebDriver to be installed on Windows.

If you're on Windows:
  1. Download EdgeDriver from: https://developer.microsoft.com/en-us/microsoft-edge/tools/webdriver/
  2. Extract msedgedriver.exe
  3. Update 'driver_path' in backend/app.py (line 36) to point to your msedgedriver.exe location

If you're on Mac/Linux:
  ⚠️ This application is designed to run on Windows with EdgeDriver.
  Please run this on your Windows office machine where EdgeDriver is configured.
"""
        print(error_msg)
        return False, None, None, {
            'error': 'EdgeDriver not found',
            'details': error_msg,
            'current_os': current_os
        }
    
    with tempfile.TemporaryDirectory() as temp_dir:
        options = webdriver.EdgeOptions()
        prefs = {
            "download.default_directory": temp_dir,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True,
        }
        options.add_experimental_option("prefs", prefs)
        # Optional: Add headless mode for server deployment
        # options.add_argument("--headless")
        # options.add_argument("--disable-gpu")

        service = Service(executable_path=driver_path)
        driver = webdriver.Edge(service=service, options=options)
        wait = WebDriverWait(driver, 20)

        try:
            # Login
            driver.get(website)
            wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="id_username"]'))).send_keys(USERNAME)
            wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="id_password"]'))).send_keys(PASSWORD)
            wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div/div/form/div/button'))).click()
            time.sleep(2)

            Comparison_button_xpath = '/html/body/div[3]/ul/li[4]/label/span'
            wait.until(EC.element_to_be_clickable((By.XPATH, Comparison_button_xpath))).click()
            time.sleep(3)

            output_data = [] 
            unique_app_names = set()
            failed_comparisons = []

            # Process each row
            for row in rows_data:
                Reference_Server = row['Reference_Server']
                Current_Server = row['Current_Server']
                Model_id_reference = row['Model_id_reference']
                Model_id_current = row['Model_id_current']
                Country_Reference = row['Country_reference']
                Country_Current = row['Country_current']
                Infolink_version_reference = row['Infolink_version_reference']
                Infolink_version_current = row['Infolink_version_current']

                # Fill form
                ref_server_xpath = '//*[@id="id_Server_Type_Reference"]'
                driver.execute_script("arguments[0].click();", wait.until(EC.element_to_be_clickable((By.XPATH, ref_server_xpath))))
                driver.find_element(By.XPATH, f"{ref_server_xpath}/option[1]" if Reference_Server.lower() == "staging" else f"{ref_server_xpath}/option[2]").click()

                cur_server_xpath = '//*[@id="id_Server_Type_Current"]'
                driver.execute_script("arguments[0].click();", wait.until(EC.element_to_be_clickable((By.XPATH, cur_server_xpath))))
                driver.find_element(By.XPATH, f"{cur_server_xpath}/option[1]" if Current_Server.lower() == "staging" else f"{cur_server_xpath}/option[2]").click()

                driver.find_element(By.XPATH, '//*[@id="id_Model_ID_Reference"]').clear()
                driver.find_element(By.XPATH, '//*[@id="id_Model_ID_Reference"]').send_keys(Model_id_reference)

                driver.find_element(By.XPATH, '//*[@id="id_Model_ID_Current"]').clear()
                driver.find_element(By.XPATH, '//*[@id="id_Model_ID_Current"]').send_keys(Model_id_current)

                driver.find_element(By.XPATH, '//*[@id="id_Country_Reference"]').clear()
                driver.find_element(By.XPATH, '//*[@id="id_Country_Reference"]').send_keys(Country_Reference)

                driver.find_element(By.XPATH, '//*[@id="id_Country_Current"]').clear()
                driver.find_element(By.XPATH, '//*[@id="id_Country_Current"]').send_keys(Country_Current)

                driver.find_element(By.XPATH, '//*[@id="id_InfoLink_Version_Reference"]').clear()
                driver.find_element(By.XPATH, '//*[@id="id_InfoLink_Version_Reference"]').send_keys(Infolink_version_reference)

                driver.find_element(By.XPATH, '//*[@id="id_InfoLink_Version_Current"]').clear()
                driver.find_element(By.XPATH, '//*[@id="id_InfoLink_Version_Current"]').send_keys(Infolink_version_current)

                driver.find_element(By.XPATH, '//*[@id="ComparissionForm"]/div[5]/center').click()
                time.sleep(5)

                # Download CSV
                try:
                    download_button = wait.until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="ComparissionTable_wrapper"]/div[1]/button[3]'))
                    )
                    driver.execute_script("arguments[0].click();", download_button)
                except Exception as e:
                    error_msg = f"Download button not found: {str(e)}"
                    print(f"❌ {error_msg}")
                    failed_comparisons.append({
                        'models': f"{Model_id_reference} vs {Model_id_current}",
                        'error': error_msg
                    })
                    continue 

                time.sleep(10)
                list_of_files = glob.glob(os.path.join(temp_dir, '*.csv'))
                if not list_of_files:
                    error_msg = "No CSV file was downloaded after clicking download button"
                    print(f"❌ {error_msg}")
                    failed_comparisons.append({
                        'models': f"{Model_id_reference} vs {Model_id_current}",
                        'error': error_msg
                    })
                    continue

                latest_file = max(list_of_files, key=os.path.getctime)
                new_filename = f"{Model_id_reference}_vs_{Model_id_current}.csv"
                new_filepath_in_temp = os.path.join(temp_dir, new_filename) 
                if os.path.basename(latest_file) != new_filename:
                    os.rename(latest_file, new_filepath_in_temp)

                output_data.append({
                    "Reference_Model": Model_id_reference,
                    "Current_Model": Model_id_current,
                    "Downloaded_File_Path": new_filepath_in_temp,
                })
                print(f"Downloaded and Saved (Temp): {new_filepath_in_temp}")
                time.sleep(10)
                
                # Collect app names - Include all apps from CSV for comparison
                with open(new_filepath_in_temp, mode='r', newline='', encoding='utf-8') as file:
                    reader = csv.DictReader(file)
                    for r in reader:
                        app_name = r.get("App Name")
                        if app_name and app_name.strip():  # Only check if app_name exists and is not empty
                            unique_app_names.add(app_name.strip())

            # Check if all comparisons failed
            if not output_data:
                error_details = "\n".join([f"  • {f['models']}: {f['error']}" for f in failed_comparisons])
                error_message = f"❌ All comparisons failed. No data was retrieved.\n\nFailed Comparisons:\n{error_details}"
                print(error_message)
                return False, None, None, {
                    'error': 'All comparisons failed',
                    'failed_comparisons': failed_comparisons,
                    'details': error_message
                }

            # Create Master Excel in temporary directory
            print(f"\n📊 Collected {len(unique_app_names)} unique app names from CSV files")
            app_names_df = pd.DataFrame(unique_app_names, columns=["App Name"])
            master_excel_file = os.path.join(temp_dir, "master_Excel.xlsx")
            app_names_df.to_excel(master_excel_file, index=False)
            print(f"Master Excel file created at: {master_excel_file}")

            master_df = pd.read_excel(master_excel_file)

            # Process CSVs while still in temp_dir
            for data_entry in output_data:
                csv_filepath = data_entry["Downloaded_File_Path"]
                csv_df = pd.read_csv(csv_filepath)

                column1 = extract_name(csv_df.columns[1])
                column2 = extract_name(csv_df.columns[2])
                model_id1 = csv_df.columns[1].split()[5]
                model_id2 = csv_df.columns[2].split()[5]
                model_name1 = model_dict.get(model_id1, model_id1)
                model_name2 = model_dict.get(model_id2, model_id2)
                new_column1 = f"{column1}_Reference_Model \n {model_name1}"
                new_column2 = f"{column2}_Current_Model \n {model_name2}"

                for app_name in master_df["App Name"]:
                    matching_rows = csv_df[csv_df["App Name"] == app_name]
                    if not matching_rows.empty:
                        extracted_data = matching_rows.iloc[:, 1:3]
                        master_df.loc[master_df["App Name"] == app_name, [new_column1, new_column2]] = extracted_data.iloc[:, 0:2].values
                    else:
                        master_df.loc[master_df["App Name"] == app_name, [new_column1, new_column2]] = ["Not Found", "Not Found"]

            reference_columns = [col for col in master_df.columns if "_Reference_Model" in col]
            current_columns = [col for col in master_df.columns if "_Current_Model" in col]
            other_columns = [col for col in master_df.columns if col not in reference_columns + current_columns]
            new_order = other_columns + reference_columns + current_columns
            master_df = master_df[new_order]

            # Compute State Value column
            print("\nComputing State Value column...")
            master_df = compute_state_value(master_df)
            
            # Display summary statistics
            state_counts = master_df['State Value'].value_counts().sort_index()
            print("\nState Value Summary:")
            for state, count in state_counts.items():
                print(f"  State Value {state}: {count} rows")

            master_df.to_excel(master_excel_file, index=False)
            print(f"Master Excel file updated with State Value: {master_excel_file}")

            # Apply row colors based on State Value
            print("\nApplying row colors...")
            apply_row_colors(master_excel_file)
            
            # Read Excel file as bytes for download
            with open(master_excel_file, 'rb') as f:
                excel_bytes = f.read()
            
            # Calculate statistics
            state_counts = master_df['State Value'].value_counts().sort_index()
            statistics = {
                'state_0_count': int(state_counts.get(0.0, 0)),
                'state_1_count': int(state_counts.get(1.0, 0)),
                'state_21_count': int(state_counts.get(2.1, 0)),
                'state_22_count': int(state_counts.get(2.2, 0)),
                'total_rows': len(master_df),
                'total_columns': len(master_df.columns)
            }
            
            # Store in global variable for access by other endpoints
            global latest_data
            latest_data['master_df'] = master_df
            latest_data['excel_bytes'] = excel_bytes
            latest_data['statistics'] = statistics

            return True, master_df, excel_bytes, statistics

        except Exception as e:
            print(f"Error in automation: {str(e)}")
            traceback.print_exc()
            raise e
        finally:
            driver.quit()

@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint."""
    return jsonify({"status": "healthy", "message": "Backend is running"})

@app.route('/api/model-names', methods=['GET'])
def get_model_names_endpoint():
    """
    Endpoint to fetch all available model names from the database.
    
    Returns:
        JSON array of model names
    """
    try:
        model_names = get_all_model_names()
        return jsonify({
            "success": True,
            "model_names": model_names,
            "count": len(model_names)
        })
    except Exception as e:
        print(f"Error fetching model names: {str(e)}")
        return jsonify({
            "success": False,
            "message": f"Failed to fetch model names: {str(e)}"
        }), 500

@app.route('/api/infolink-servers', methods=['GET'])
def get_infolink_servers_endpoint():
    """
    Endpoint to fetch all available InfoLink servers from the database.
    
    Query Parameters:
        server_type (optional): Filter by 'current', 'reference', or 'both' (default: all)
        active_only (optional): If 'true', only return active servers (default: true)
    
    Returns:
        JSON array of InfoLink server names
    """
    try:
        server_type = request.args.get('server_type', None)
        active_only = request.args.get('active_only', 'true').lower() == 'true'
        
        servers = get_all_infolink_servers(server_type, active_only)
        return jsonify({
            "success": True,
            "servers": servers,
            "count": len(servers)
        })
    except Exception as e:
        print(f"Error fetching InfoLink servers: {str(e)}")
        return jsonify({
            "success": False,
            "message": f"Failed to fetch InfoLink servers: {str(e)}"
        }), 500

@app.route('/api/countries', methods=['GET'])
def get_countries_endpoint():
    """
    Endpoint to fetch all available countries from the database.
    
    Query Parameters:
        country_type (optional): Filter by 'current', 'reference', or 'both' (default: all)
        active_only (optional): If 'true', only return active countries (default: true)
    
    Returns:
        JSON array of country names
    """
    try:
        country_type = request.args.get('country_type', None)
        active_only = request.args.get('active_only', 'true').lower() == 'true'
        
        countries = get_all_countries(country_type, active_only)
        return jsonify({
            "success": True,
            "countries": countries,
            "count": len(countries)
        })
    except Exception as e:
        print(f"Error fetching countries: {str(e)}")
        return jsonify({
            "success": False,
            "message": f"Failed to fetch countries: {str(e)}"
        }), 500

@app.route('/api/run-automation', methods=['POST'])
def run_automation_endpoint():
    """
    Endpoint to run the automation process.
    
    Expected JSON payload:
    {
        "rows": [
            {
                "Reference_Server": "Staging",
                "Current_Server": "Production",
                "Model_name_reference": "G95SC",
                "Model_name_current": "G95SD",
                "Country_reference": "US",
                "Country_current": "US",
                "Infolink_version_reference": "1.0",
                "Infolink_version_current": "2.0"
            }
        ]
    }
    """
    try:
        data = request.get_json()
        
        if not data or 'rows' not in data:
            return jsonify({
                "success": False,
                "message": "Invalid request. 'rows' field is required."
            }), 400
        
        rows = data['rows']
        
        if not rows or len(rows) == 0:
            return jsonify({
                "success": False,
                "message": "At least one row is required."
            }), 400
        
        # Validate each row and convert model names to model IDs
        required_fields = [
            'Reference_Server', 'Current_Server',
            'Model_name_reference', 'Model_name_current',
            'Country_reference', 'Country_current',
            'Infolink_version_reference', 'Infolink_version_current'
        ]
        
        processed_rows = []
        for idx, row in enumerate(rows):
            # Validate all fields are present
            for field in required_fields:
                if field not in row or not row[field]:
                    return jsonify({
                        "success": False,
                        "message": f"Row {idx + 1}: Missing or empty field '{field}'"
                    }), 400
            
            # Convert model names to model IDs
            model_id_ref = get_model_id_from_name(row['Model_name_reference'])
            model_id_curr = get_model_id_from_name(row['Model_name_current'])
            
            if not model_id_ref:
                return jsonify({
                    "success": False,
                    "message": f"Row {idx + 1}: Invalid Model Name Reference '{row['Model_name_reference']}'"
                }), 400
            
            if not model_id_curr:
                return jsonify({
                    "success": False,
                    "message": f"Row {idx + 1}: Invalid Model Name Current '{row['Model_name_current']}'"
                }), 400
            
            # Create processed row with model IDs
            processed_row = {
                'Reference_Server': row['Reference_Server'],
                'Current_Server': row['Current_Server'],
                'Model_id_reference': model_id_ref,
                'Model_id_current': model_id_curr,
                'Country_reference': row['Country_reference'],
                'Country_current': row['Country_current'],
                'Infolink_version_reference': row['Infolink_version_reference'],
                'Infolink_version_current': row['Infolink_version_current']
            }
            processed_rows.append(processed_row)
        
        # Run the automation with converted model IDs
        result = run_automation(processed_rows)
        
        # Handle tuple unpacking based on success
        if len(result) == 4:
            success, master_df, excel_bytes, statistics = result
        else:
            # Unexpected return format
            return jsonify({
                "success": False,
                "message": "Unexpected automation result format"
            }), 500
        
        if success:
            # Convert DataFrame to JSON-serializable format with color information
            data_dict = master_df.to_dict('records')
            
            # Add color information to each row
            for row in data_dict:
                state_value = row.get('State Value')
                row['row_color'] = get_row_color(state_value)
                # Format State Value to show 2.1 instead of 2.1000
                if pd.notna(state_value):
                    row['State Value'] = float(state_value)
            
            # Convert column names to list
            columns = master_df.columns.tolist()
            
            return jsonify({
                "success": True,
                "message": "Automation completed successfully. Master Excel file has been created.",
                "data": data_dict,
                "columns": columns,
                "statistics": statistics
            })
        else:
            # When success is False, statistics contains error information
            error_info = statistics if isinstance(statistics, dict) else {}
            error_message = error_info.get('details', 'Automation failed.')
            
            return jsonify({
                "success": False,
                "message": error_message,
                "error_details": error_info.get('failed_comparisons', [])
            }), 400
            
    except Exception as e:
        print(f"Error in endpoint: {str(e)}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "message": f"An error occurred: {str(e)}"
        }), 500

@app.route('/api/download-excel', methods=['GET'])
def download_excel():
    """
    Endpoint to download the generated Master Excel file.
    
    Returns:
        Excel file as attachment
    """
    try:
        if latest_data['excel_bytes'] is None:
            return jsonify({
                "success": False,
                "message": "No Excel file available. Please run automation first."
            }), 404
        
        # Create a BytesIO object from the stored bytes
        excel_io = io.BytesIO(latest_data['excel_bytes'])
        excel_io.seek(0)
        
        return send_file(
            excel_io,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='master_Excel.xlsx'
        )
    except Exception as e:
        print(f"Error downloading Excel: {str(e)}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "message": f"Failed to download Excel file: {str(e)}"
        }), 500


@app.route('/api/get-data', methods=['GET'])
def get_data():
    """
    Endpoint to retrieve the processed data without re-running automation.
    
    Returns:
        JSON with data, columns, and statistics
    """
    try:
        if latest_data['master_df'] is None:
            return jsonify({
                "success": False,
                "message": "No data available. Please run automation first."
            }), 404
        
        master_df = latest_data['master_df']
        statistics = latest_data['statistics']
        
        # Convert DataFrame to JSON-serializable format with color information
        data_dict = master_df.to_dict('records')
        
        # Add color information to each row
        for row in data_dict:
            state_value = row.get('State Value')
            row['row_color'] = get_row_color(state_value)
            # Format State Value
            if pd.notna(state_value):
                row['State Value'] = float(state_value)
        
        # Convert column names to list
        columns = master_df.columns.tolist()
        
        return jsonify({
            "success": True,
            "data": data_dict,
            "columns": columns,
            "statistics": statistics
        })
    except Exception as e:
        print(f"Error retrieving data: {str(e)}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "message": f"Failed to retrieve data: {str(e)}"
        }), 500


if __name__ == '__main__':
    print("Starting Flask server...")
    print(f"Backend API running on http://localhost:5001")
    app.run(debug=True, host='0.0.0.0', port=5001)
